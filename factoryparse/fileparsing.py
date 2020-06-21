import sys
import json
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import os
import xlrd as xl
import xlwt 
from xlwt import Workbook 
from xlutils.copy import copy
from openpyxl import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.styles.colors import Color

basepath = os.path.dirname(os.path.realpath(__file__));
sysargv_path = sys.argv[0];
filename = sys.argv[1];

full_file_path = os.path.dirname(os.path.realpath(__file__)) + '/uploads/' + sys.argv[1];
if(os.path.exists(full_file_path)) :
    excel_file = full_file_path
    Obj_excelread = pd.read_excel(excel_file);
    header_excel = Obj_excelread.head();
else :
	print('\n ########## Not exists ########## \n ');

wb=load_workbook(full_file_path)
sheet=wb.active
max_row=sheet.max_row
max_column=sheet.max_column

for i in range(2,max_row+1):
    for j in range(1,max_column+1):
        cell_obj=sheet.cell(row=i,column=j)
        if cell_obj.value != 'FLEXPORT' and j == 2:
        	print('Val = ',cell_obj.value)
	        print('I = ',i)
	        print('J = ',j)
	        sheet.cell(row=i, column=max_column).value = 'Invalid Customer'
	        sheet.cell(row=i, column=max_column).font = Font(color = "FF0000");

wb.save(full_file_path)