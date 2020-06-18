import sys
import json
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import os
import xlrd as xl
import xlwt 
from xlwt import Workbook 
from xlutils.copy import copy
from openpyxl import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.styles.colors import Color

basepath = ' \n ====== Base Path ======= ' + os.path.dirname(os.path.realpath(__file__))  + ' ----------- \n ' ;
print(basepath);
print('\n');

sysargv_path = ' \n  ====== sysargv_path 0 ======= '  + sys.argv[0] + ' ----------- \n ';
print(sysargv_path);
print('\n');


filename = ' \n ====== sysargv_path 1 ======= '  + sys.argv[1] + ' ----------- \n ';
print(filename);
print('\n');

arr = {
  "filename": "OFR",
  "columns": 30,
  "from": "Online"
}
encodejson = json.dumps(arr);

print('\n');
print(encodejson);

full_file_path = os.path.dirname(os.path.realpath(__file__)) + '/uploads/' + sys.argv[1];
if(os.path.exists(full_file_path)) :
	print('\n $$$$$$$$$$ File exists $$$$$$$$$$ \n ');
	print('\n ^^^^^^^^^^ ' + full_file_path + ' ^^^^^^^^^^ \n ');
else :
	print('\n ########## Not exists ########## \n ');

excel_file = full_file_path
Obj_excelread = pd.read_excel(excel_file);
header_excel = Obj_excelread.head();
#print('\n Excel Header');
#print(header_excel);
#print('\n Excel Reader');
#print(Obj_excelread);


# Create New sheet Start
# d1 = pd.DataFrame({"A":['Bob','Joe', 'Mark'], 
#                "B":['5', '10', '20']})
# d2 = pd.DataFrame({"A":['Jeffrey','Ann', 'Sue'], 
#                 "B":['1', '2', '3']})
# dfs = [d1,d2]
# for i in range(len(dfs)):
#     sheet = 'd'+str(i+1)
#     data = dfs[i]
#     writer = pd.ExcelWriter(full_file_path,engine='openpyxl', mode='a')
#     writer.book = load_workbook(full_file_path)
#     data.to_excel(writer,sheet_name=sheet)
#     #data.to_excel(writer)
#     writer.save()
#     writer.close()
# Create New sheet End


#Added Error Column at the end of the excel sheet while validating the file
#Total no of rows and columns using the xlrd library
loc = (full_file_path)
wb = xl.open_workbook(loc)
s1 = wb.sheet_by_index(0)
s1.cell_value(0,0)  
print("No. of rows:", s1.nrows)
print("No. of columns:", s1.ncols)

def add_column(sheet_name, column):
    ws = wb[sheet_name]
    new_column = s1.ncols + 1
    new_row = ws.max_row + 1

    mxcl = ws.max_column;
    mxrw = ws.max_row;

    print('ws.max_column = ');
    print(mxcl);
    print('ws.max_row = ');
    print(mxrw);

    for rowy, value in enumerate(column, start=1):
        ws.cell(row=rowy, column=new_column, value=value)

wb = openpyxl.load_workbook(full_file_path)
add_column('ofr', ['Error'])
wb.save(full_file_path)


#Read Whole excel file using xlrd - Start
# loc = (full_file_path)
# wb = xl.open_workbook(loc)
# sh = wb.sheet_by_index(0)

# newwb = copy(wb)
# sheet1 = newwb.get_sheet(0)
# mxcl = sh.ncols;

# for i in range(sh.nrows):
#  for j in range(sh.ncols):
#  	 if sh.cell_value(i,j) != 'FLEXPORT' and j == 1:
#  	 	print('If============',sh.cell_value(i,j))
#  	 	sheet1.write(i, mxcl, 'Invalid Customer')
 	 	
# newwb.save(full_file_path)
#Read Whole excel file using xlrd - End



#Read and Write using the openpyxl - Start
wb=load_workbook(full_file_path)
sheet=wb.active
max_row=sheet.max_row
max_column=sheet.max_column
print('NEW MAX ROW ',max_row);
print('NEW MAX COLUMN ',max_column);

# def InvalidCustomer(rownum, columnnum):
# 	Color.RED = 'FFFF0000';
# 	_cell = sheet.cell(row=rownum, column=columnnum)
# 	_cell.style.font.color.index = Color.RED
# 	_cell.style.font.name = 'Arial'
# 	_cell.style.font.size = 12
# 	_cell.style.font.bold = True
# 	_cell.style.alignment.wrap_text = True
# 	_cell.style.fill.fill_type = Fill.FILL_SOLID
# 	_cell.style.fill.start_color.index = Color.DARKRED

for i in range(2,max_row+1):
    for j in range(1,max_column+1):
        cell_obj=sheet.cell(row=i,column=j)
        if cell_obj.value != 'FLEXPORT' and j == 2:
        	print('Val = ',cell_obj.value)
	        print('I = ',i)
	        print('J = ',j)
	        sheet.cell(row=i, column=max_column).value = 'Invalid Customer'
	        sheet.cell(row=i, column=max_column).font = Font(color = "FF0000");
	        # sheet.cell(row=i, column=max_column).font.color = 'FFFF0000'
	        # sheet.cell(row=i, column=max_column).font.name = 'Verdana'
	        # sheet.cell(row=i, column=max_column).font.size = 15
	        # sheet.cell(row=i, column=max_column).font.bold = True
	        
	        #InvalidCustomer(i, j)
wb.save(full_file_path)

        #print(cell_obj.value,end=' | ')

#Read and Write using the openpyxl - End

