import sys
import json
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import os
import xlrd as xl
import xlwt 
from xlwt import Workbook 
from xlutils.copy import copy
from openpyxl import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.styles.colors import Color

basepath = os.path.dirname(os.path.realpath(__file__));
sysargv_path = sys.argv[0];
filename = sys.argv[1];

full_file_path = os.path.dirname(os.path.realpath(__file__)) + '/uploads/' + sys.argv[1];

if(os.path.exists(full_file_path)) :
    excel_file = full_file_path
    Obj_excelread = pd.read_excel(excel_file);
    header_excel = Obj_excelread.head();
else :
	print('Error');

print('full_file_path');
print(full_file_path);

wb=load_workbook(full_file_path)
sheet=wb.active
max_row=sheet.max_row
max_column=sheet.max_column
addcolumn = 0;

aCustomers = ['ACCURATETEST','ACUSTOMFL','AERONET','AFCWW','AGILITY','AHARTRODT','AITFREIGHT','ALEXANDER','ALLFREI.BE','ALPI','AMEXPORT','APC','APLLOG','APL_TEST','ATCOLUWIE','ATLANTICAP','AUMUND.BE','BDG','BDP','BEANCOANR','bebentmer','BECATEGRM','becrgoanr','BEDENYREK','BEELITANR','BEESSEWIL','BEFASTANR','BEHAMAWET','BEILBANR','BEJANSGEN','bemanuanr','BEPARTDEU','beremagri','BEREMGANR','BESDKANR','BESPARANR','BESTEIANR','BETOSANR','BETRANOVE','beupsanr','bezieglog','BIZCON','BLUEANCHO2','BLUEANCHOR','BLUECARG','BLUEH2O','BOLLORE','BOSCHMANST','BOSMAN','BTXUSA','CARCA','CARGOLINKI','CARGOPARTNER','CATAPULT','CCLOG','CDS','CEVA','CHALLENGE','CHHOFBAS','CHROBINSON','CHSOMAGVA','CHTRANBER','CLASQUIN','CLEARFR','COMBLOGNET','COMPASS','CONCERTGRP','CRANEWW','CTT','DACHSER','DAMCO','DECARGDOR','DEFRMEMAN','DEINFODUI','DEINFONIE','denycargo','DERINGER','DESTLSHAR','DFLCFL','DHL','DHXUSA','DISSACO','DSV','DSV2','DSVDOWC','DSVOWN','EASTERSHIP','EMBASSY','EMBATL','EMOTRANS','ESHIPP','ESTESFW','EUMED','EXPEDITORS','EXTRALOG','FEDEXTRADE','FLEXPORT','foodcareplus','FRACHTFWO','frbansaent','frcomeneu','FRFRACENT','FRHEPPLSQ','FRINTERIX','FRITSLESQ','FRMARQTET','FRNDOLESQ','frrohlent','FRSAGAHEI','FRTHUB','FRTRIGHT','frwinstra','GBLDNEXPSTD','GEBRUDER','GEFCOAPEN','GEODIS','GESUSA','GLCDSLA','GLINER','GLOBALCONT','GONDRAND','GONRAND','GREENCAR','HANJINTH','HANKYU','HELLMANN','HELLMANN-CNH','HELLMANN-FEMO','HHGPLUS','HHGSTD','HHGVID','HITACHITH','HONOURLL','HORINTCAR','ICONTAINER','IJSGLOBAL','INTELLSCM','INTRAMAR','JAS-PENTAIR','JASOCEAN','JAS_TEST','JOHNSCON','katoenapen','KDL','KERRY','KERRYFR','KINTETSU','KLINE','KNOWN','KNTRUCK','LANDSTAR','LAUFER','LEMANWW','LESCHACO','LFLOG','LOGAIR','LOGFRET','LOGWIN','LYNDEN','MARISOL','MARS','MAURICE','MDCTEST','MERIDIAN','MGMAHER','MICOR','MIDAMERICA','MIQ','MITSUISOKO','MOHAWK','MOLLOG','MORRISONEXPRESS','MOTXX','MRFORTH','NACRFQ','NATCOZUR','NDO','NEWWAVE','NFIGLOBAL','NIPPON','NLBESTRTM','NLCOASTIL','NNR','NOATUM','NOMINATIONS','OCEANGLOB','OEC','OHLBARTHCO','OIAGLOBAL','OLDOMIFRE','PACINTL','PANALOGS','PANALPINA','PAOWN','PENTAGON','PERILN','PIXLINE','POWCON','PRIVATETESTCUST','QUALITYFRE','RADIANTCON','RAM','remant','RHENUS','RHENUSLOG','RIM','ROAR','ROHLIG','ROSEC','ROYALCL','RRDILN','SAVINODEL','SCANSHIP','SCARBROUGH','SCHENKER','SDV','SEKO','SENATORINT','SHAPIRO','SILVER','SPEEDMARK','STANDARD','STEAM','STUTE','SUPCHAIN','TAGGAMO','TENDER','TEST','TESTACCURATE','TESTPRIVATE','TESTPRIVATE','TEST_JAS','TEST_UPLOAD','THBOLTS','THSTD','TIGERS','TOLIM','TOLL','TPOFLA','TRAGLOBAL','TRANSGROUP','TRANSOCANT','TRANSWORLD','TRISTARFR','TTSWORLD','UNIFORWARD','UNION','UNIQUELOG','UNIQUNY','UPS','USWOKY','UTC','VANTECHITACHI','VINWORLD','WEISSROHLIG','WERNERGLB','WERNNE','WFA','WOODLAND','XCELL','XPO','YUSEN','YusenCatapult','YUSENUSA','ZENLHR','ZIEGLER'];
aMembers = ['11DX','11QU','11WJ','12','12GE','12PD','12UY','13DF','2222','22AA','3','ACL','ACLU','ADMU','AELQ','AFPN','AFSL','AGLY','AGSW','ALIU','ALLC','ALPS','ALTI','AMIG','AMLU','ANLC','ANLU','ANNU','ANRM','ANZD','APLL','APLU','AQUA','ARKU','ASRI','ATMI','AWPC','BAXU','BDNU','BGFU','BJCH','BLPL','BNUH','BORU','BRGU','BROU','BTTM','BVBZ','BVSN','BWSS','CAFS','CALP','CAMN','CANU','CASU','CBOG','CBSP','CCLL','CCNR','CCXL','CCZU','CDOD','CFRF','CHAK','CHHK','CHIW','CHNJ','CHRD','CHSL','CHSS','CHTR','CHVW','CJFF','CKL','CLAM','CLIB','CLPL','CMAI','CMAU','CMDD','CMDU','CMNU','CNCU','CNIU','COAR','COHE','COLL','CONS','COSS','COSU','CPLB','CPSP','CPSU','CRAU','CRES','CSAV','CSCL','CTLT','CTNU','DAAE','DACH','DALL','DALU','DAMC','DAYU','DCHR','DFDS','DHLC','DJSC','DOLQ','DOMI','DWIC','EASH','EASL','ECCT','ECIU','ECON','EGLV','EGOI','EIMU','EISU','ELAN','ELDA','ELGS','EMES','EQLI','ESPU','EUCU','EVHA','EVMC','EVRG','FANE','FCLC','FEDX','FLNV','FMPL','FORTU','FRCN','FRLN','FSYL','FTNX','FYCN','GAGP','GDSL','GEKU','GETU','GFAL','GFSC','GLSL','GMBD','GMGO','GOSU','GRIU','GWFT','GWLY','GWUI','HABU','HAHU','HALU','HDMU','HHGD','HHSU','HJSC','HLCU','HLCV','HOEG','HRZD','HTML','HUBL','HWWL','HYBU','HZOM','IALU','ICCP','IICC','INDC','INXL','IPII','ISMU','ITAU','ITEL','ITMA','IWTL','JASF','JASO','JINC','KANW','KINO','KKLU','KLNE','KMTC','KMTU','KOSL','LMCU','LOGS','LTNV','LYKL','LYKU','MAEU','MANS','MATS','MAWI','MCAW','MCLP','MCPU','MCSM','MDCQA','MECU','MEDU','MESC','MFTU','MFUS','MISC','MISU','MMMA','MOLU','MRUB','MSAJ','MSCU','MSSG','MTDK','MTMK','MULT','MXLU','NAQA','NCLL','NIDU','NIVL','NLSA','NODA','NOSU','NOVL','NPLX','NSAU','NSSL','NWLU','NYKS','NYKS','OBLU','OCCO','OMEG','ONEY','OOLU','OPDR','OPDU','OSLS','OXCO','PAAU','PABV','PACC','PACS','PAMA','PCIU','PCLL','PCLU','PCSC','PCSL','PELS','PENE','PGLN','PGPR','PILN','PILU','PIXL','PKFC','PLLU','PMLU','PNEP','PNLP','POBO','POBU','POCL','POLN','POTA','PROF','PSHY','PSLL','PYRD','PYRR','QMNS','QNCH','RAIL','REGU','RELP','RELU','RIMH','RLUS','RSAL','RWSU','SAFM','SAXC','SBIS','SCAN','SCGL','SCHK','SCIU','SCJU','SCKO','SCPH','SCSH','SEAM','SEAN','SEAU','SEFN','SEJJ','SENU','SENV','SESA','SHKK','SHPT','SIKU','SISJ','SITC','SITS','SKLU','SMLM','SMLU','SNKO','SNTU','SOPH','SREE','SSAE','SSAR','SSAT','SSBO','SSCC','SSCH','SSCL','SSCY','SSCZ','SSEC','SSFR','SSHF','SSHU','SSLH','SSLL','SSPE','SSPY','SSRL','SSSK','SSTR','SSUY','STAU','STIA','SUDU','TCIJ','TCKU','TCSB','TCUG','te123','tes','Tes1','Tes12','Tes13','Tes4','Tes6','Tes8','TEST','TEST1','TGHF','TIIP','TJFC','TLPL','TLWN','TMGB','TMGT','TMSD','TORU','TRAL','TRAN','TRBR','TRCK','TRGY','TRKU','TRNH','TSCW','TSGL','TSLL','TWDP','U107','U108','U109','U110','UAFL','UASC','UASU','UBCU','UNTD','UPSS','USLB','UTIW','VECT','VLSJ','WDND','WECU','WHLC','WLWH','WSS','WTLU','WWAL','WWAT1','WWLP','x1','XHSL','XXXX','YMLU','Z030','Z034','Z037','Z038','Z043','Z046','Z048','Z051','ZCSG','ZIMU','ZLIN'];

for i in range(2,max_row+1):
    for j in range(1,max_column+1):
        cell_obj=sheet.cell(row=i,column=j);
        err_cell_val=sheet.cell(row=i, column=max_column);
        if cell_obj.value not in aCustomers and j == 1:
	        sheet.cell(row=i, column=max_column).value = str(sheet.cell(row=i, column=max_column).value) + str('Invalid Customer');
	        sheet.cell(row=i, column=max_column).font = Font(color = "FF0000");
        if cell_obj.value not in aMembers and j == 2:
            sheet.cell(row=i, column=max_column).value = str(sheet.cell(row=i, column=max_column).value) + str('Invalid Member');
            sheet.cell(row=i, column=max_column).font = Font(color = "FF0000");
wb.save(full_file_path)
print("Done with File Full Path = ", full_file_path);